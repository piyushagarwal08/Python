import openpyxl as xl
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import sys

root = tk.Tk()
root.withdraw()

# Greetings
messagebox.showinfo(title="Greetings",message="Welcome from PapTech Engineers & Associates")

#Ask for File Input
while True:
    messagebox.showinfo(title="Input File",message="Please Select the Technibrite Reading file with .xlsx extension")
    file_path = filedialog.askopenfilename()
    if ".xlsx" in file_path:
        break

# if no file is selected then exit the program
if file_path is None or len(file_path) == 0:
    messagebox.showinfo("File Error","Please Select an Proper Excel file")
    sys.exit(1)

# Create a workbook variable    
workbook = xl.load_workbook(file_path)

sheet1 = workbook.active # sheet object
row = sheet1.max_row  # total no of rows
col = sheet1.max_column  # total no of rows

# Find column number of X,Y and Z
for i in range(1,col+1):
    if sheet1.cell(1,i).value == "X":
        x_col = i
    elif sheet1.cell(1,i).value == "Y":
        y_col = i
    elif sheet1.cell(1,i).value == "Z":
        z_col = i
    else:
        continue
if x_col is None or y_col is None or z_col is None:
    messagebox.showerror("File Header Error","Selected file should have X,Y and Z as headers")
    sys.exit(1)

# Typing Output File Headers

sheet1.cell(1,z_col+1).value = "L"
sheet1.cell(1,z_col+2).value = "A"
sheet1.cell(1,z_col+3).value = "B"
sheet1.cell(1,z_col+4).value = "RX"
sheet1.cell(1,z_col+5).value = "RY"
sheet1.cell(1,z_col+6).value = "RZ"


# For each entry
for i in range(2,row+1):
    if sheet1.cell(i,x_col).value is None or sheet1.cell(i,y_col).value is None or sheet1.cell(i,z_col).value is None:
        continue
    X = float(sheet1.cell(i,x_col).value)
    Y = float(sheet1.cell(i,y_col).value)
    Z = float(sheet1.cell(i,z_col).value)
    L = 10 * (Y)**0.5
    A = 17.5 * (1.02*X - Y) / (Y ** 0.5)
    B = 7 * (Y - 0.847 * X) / (Y**0.5)
    RX = X/(X+Y+Z)
    RY = Y / (X+Y+Z)
    RZ = Z/(X+Y+Z)
    sheet1.cell(i,z_col+1).value = L
    sheet1.cell(i,z_col+2).value = A
    sheet1.cell(i,z_col+3).value = B
    sheet1.cell(i,z_col+4).value = RX
    sheet1.cell(i,z_col+5).value = RY
    sheet1.cell(i,z_col+6).value = RZ

workbook.save(file_path)
